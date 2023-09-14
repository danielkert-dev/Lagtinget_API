import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# User-Agent header
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
}

#@ API endpoints
seatings_url = "https://api.lagtinget.ax/api/seatings.json"
specific_seating_base_url = "https://api.lagtinget.ax/api/seatings"

# Details

presence_states_url = "https://api.lagtinget.ax/api/presence_states.json"
theme_url = "https://api.lagtinget.ax/api/themes.json"
person_url = "https://api.lagtinget.ax/api/persons/"

#@ Fetch data from the seating list API endpoint
seatings_response = requests.get(seatings_url, headers=headers)

# Check if the request was successful
if seatings_response.status_code == 200:
    seatings_data = seatings_response.json()
else:
    print(f"Failed to fetch seating list data. Status code: {seatings_response.status_code}")
    seatings_data = None


#!-------------------------------------------------------------------------------------------------------------------------------------!#


#@ Choose a specific seating ID from the list (you can modify this as needed)
specific_seating_id = "54501"#seatings_data[20]["id"]

specific_seating_date = "2023-09-06 13:51:08"

# Find the seating with the matching ID or date
specific_seating_data = None
for seating in seatings_data:
    if seating["id"] == specific_seating_id or seating["date"] == specific_seating_date:
        specific_seating_data = seating["id"]
        break

#? Fetch data for the specific seating
specific_seating_response = requests.get(f"{specific_seating_base_url}/{specific_seating_data}.json", headers=headers)

# Check if the request was successful
if specific_seating_response.status_code == 200:
    specific_seating_data = specific_seating_response.json()
else:
    print(f"Failed to fetch specific seating data. Status code: {specific_seating_response.status_code}")
    specific_seating_data = None

#@ Create a new Excel workbook
excel_file_name = "seating_data.xlsx"
workbook = Workbook()

#! Seating data

plenum_df = pd.DataFrame(seatings_data)

# Add item data to the Excel workbook in a new sheet named 'Items'
item_sheet = workbook.create_sheet(title='Plenum')
for row in dataframe_to_rows(plenum_df, index=False, header=True):
    item_sheet.append(row)


#! Extract the relevant information for the 'Info' sheet
if specific_seating_data is not None:
    seating_id = specific_seating_data['id']
    title = specific_seating_data['title']
    date = specific_seating_data['date']
    
    info_data = {'seating_id': [seating_id], 'title': [title], 'date': [date]}
    info_df = pd.DataFrame(info_data)

    # Add info data to the Excel workbook in a new sheet named 'Info'
    info_sheet = workbook.create_sheet(title='Info', index=0)
    for row in dataframe_to_rows(info_df, index=False, header=True):
        info_sheet.append(row)

    # Specify the desired column widths for the 'Info' sheet
    info_column_widths = {'A': 10, 'B': 35, 'C': 18}

    # Set the column widths for the 'Info' sheet
    info_sheet = workbook['Info']
    for column, width in info_column_widths.items():
        info_sheet.column_dimensions[column].width = width


#! Extract the relevant information for the 'Items' sheet
if specific_seating_data is not None:
    items = specific_seating_data.get('items', [])  # Handle missing 'items' key
    attendants = specific_seating_data.get('attendants', [])  # Handle missing 'attendants' key

    # Create a DataFrame for item data
    item_data = {
        'item_id': [item['id'] for item in items],
        'item_title': [item['title'] for item in items],
        'item_source': [item['source'] for item in items]
    }
    item_df = pd.DataFrame(item_data)

# Add item data to the Excel workbook in a new sheet named 'Items'
item_sheet = workbook.create_sheet(title='Items')
for row in dataframe_to_rows(item_df, index=False, header=True):
    item_sheet.append(row)


#! Extract the relevant information for the 'Attendants' sheet
# Filter out attendants with missing values

    attendant_data = {
        'person': [attendant['person'] for attendant in attendants],
        'name': [attendant['name'] for attendant in attendants],
        'type': [attendant['type'] for attendant in attendants],
        'reason': [attendant['reason'] for attendant in attendants],
        'row': [attendant['place']['row'] for attendant in attendants],
        'seat': [attendant['place']['seat'] for attendant in attendants],
        'type_title': []  # Initialize the 'type_title' key as an empty list
    }

    #! Data about present people

# Create a dictionary to store the presence state titles
presence_state_titles = {}

# Fetch presence states data
presence_states_response = requests.get(presence_states_url, headers=headers)

# Check if the request was successful
if presence_states_response.status_code == 200:
    presence_states_data = presence_states_response.json()
    # Build a dictionary of presence state titles using the presence state ID as the key
    for state in presence_states_data:
        presence_state_titles[state['id']] = state['title']
else:
    presence_states_data = None
    print(f"Failed to fetch presence states data. Status code: {presence_states_response.status_code}")

# Iterate through each attendant and fetch the corresponding type title
for attendant in attendant_data['type']:
    type_id = attendant
    type_title = presence_state_titles.get(type_id, '')
    # Add type title to the attendant_data dictionary
    attendant_data['type_title'].append(type_title)

attendant_df = pd.DataFrame(attendant_data)

# Add attendant data to the Excel workbook in a new sheet named 'Attendants'
attendant_sheet = workbook.create_sheet(title='Attendants')
for row in dataframe_to_rows(attendant_df, index=False, header=True):
    attendant_sheet.append(row)


#! Data about the speeches

person_speeches = {}


# Iterate through each person in the attendant data
for person_id in attendant_data['person']:
    # Fetch person speeches data
    person_speeches_response = requests.get(f"{person_url}{person_id}/speeches.json", headers=headers)

    # Check if the request was successful
    if person_speeches_response.status_code == 200:
        person_speeches_data = person_speeches_response.json()
        person_speeches[person_id] = person_speeches_data
    else:
        person_speeches[person_id] = None
        print(f"Failed to fetch speeches data for person ID {person_id}. Status code: {person_speeches_response.status_code}")

# Create a new sheet for speeches in the Excel workbook
speeches_sheet = workbook.create_sheet(title='Speeches')

# Iterate through each person in the attendant data
for person_id in attendant_data['person']:
    speeches_data = person_speeches[person_id]  # Get speeches data for the current person

    if speeches_data is not None:
        speeches_df = pd.DataFrame(speeches_data)
        
        # Filter speeches based on the relevant event/item_id
        relevant_speeches_df = speeches_df[speeches_df['event'].isin(item_data['item_id'])]

        # Add speeches data to the 'Speeches' sheet in the Excel workbook
        for row in dataframe_to_rows(relevant_speeches_df, index=False, header=True):
            speeches_sheet.append(row)
    else:
        # If speeches data is missing, add a placeholder row with the person ID
        speeches_sheet.append([person_id])

#! Data about the people



#!-------------------------------------------------------------------------------------------------------------------------------------!#



# Remove the default sheet created by openpyxl (Sheet)
workbook.remove(workbook['Sheet'])

# Save the Excel workbook
workbook.save(excel_file_name)
print(f"Excel file '{excel_file_name}' has been created with separate sheets for Info, Items, and Attendants.")